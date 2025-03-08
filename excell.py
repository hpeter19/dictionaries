import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')
sheet= wb['Sheet1']
cell = sheet ['b2']


for raw in range(1,sheet.max_row + 1):
    cell=sheet.cell(raw,3)
    #correcting the prices you want to modify
    corrected_price = cell.value * 0.9
#printing values of each cell 