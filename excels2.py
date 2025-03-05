#A function to handle excel workbooks
import openpyxl as xl


def process_workbook(filename):
        wb = xl.load_workbook('filename')
        sheet = wb['Sheet1']


        # Iterate over the rows
        for row in range(1, sheet.max_row + 1):
            # Access the cell in column 3 (C column)
            cell = sheet.cell(row, 3)

            # Ensure the value is numeric before multiplying
            if isinstance(cell.value, (int, float)):
                corrected_price = cell.value * 0.9

                # Access the cell in column 4 (D column) to store the corrected price
                corrected_price_cell = sheet.cell(row, 4)
                corrected_price_cell.value = corrected_price
            else:
                print(f"Non-numeric value found in row {row}, column 3.")

        # Save the modified workbook
        wb.save('filename') #replace filename with the name of your transaction
        # PRINTING NEW WORKBOOK,WITH NEW VALUES IN THE ROW
        # the name f the workbook you can choose of your choice.

